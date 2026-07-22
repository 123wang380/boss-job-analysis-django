from myApp.models import JobInfo
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import datetime


def export_jobs_to_excel():
    """导出所有职位数据为Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '招聘数据'

    # 表头
    headers = ['序号', '职位名称', '公司名称', '工作城市', '行政区', '职位类型',
               '学历要求', '工作经验', '薪资', '年终奖', '技能标签', '公司福利',
               '公司性质', '融资状态', '公司规模', '实习', '发布日期']

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(vertical='center')

    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写数据
    jobs = JobInfo.objects.all().order_by('-createTime')
    for row_idx, job in enumerate(jobs, 2):
        try:
            salary_data = json.loads(job.salary)
            if job.pratice == 0 and len(salary_data) > 1:
                salary_str = f"{salary_data[1] / 1000:.1f}K"
            else:
                salary_str = str(salary_data[0]) if salary_data else ''
        except:
            salary_str = job.salary

        try:
            work_tag = ' / '.join(json.loads(job.workTag))
        except:
            work_tag = job.workTag

        try:
            if job.companyTags != '无' and job.companyTags:
                company_tags = json.loads(job.companyTags)[0]
            else:
                company_tags = '无'
        except:
            company_tags = job.companyTags

        try:
            people_data = json.loads(job.companyPeople)
            if people_data[1] >= 10000:
                people_str = '10000人以上'
            else:
                people_str = f'{people_data[0]}-{people_data[1]}人'
        except:
            people_str = job.companyPeople

        row_data = [
            row_idx - 1, job.title, job.companyTitle, job.address, job.dist,
            job.type, job.educational, job.workExperience, salary_str,
            job.salaryMonth + '薪' if job.salaryMonth else '', work_tag,
            company_tags, job.companyNature, job.companyStatus, people_str,
            '是' if job.pratice else '否', str(job.createTime)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = thin_border

    # 调整列宽
    col_widths = [6, 20, 22, 10, 8, 12, 10, 12, 10, 8, 30, 25, 10, 10, 12, 6, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
